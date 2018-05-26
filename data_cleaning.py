__author__ = "* * treaser 2018 * * "

'''
requirements:
1、空信息，整行删除 : len(news) < 50
2、错误格式信息，如只有 版面 或者 编辑信息 等，往往会几行连在一起都是错误格式信心，整行一起删除 ：len(news) < 50
3、广告信息，可直接删除 : keywords: 公司 产品名称？

4、非杭州新闻内容，北京讯、南京讯等直接可删除，点明湖州、温州等其他地址的稿件一般情况都可直接剔除 : keywords: 全国城市名称。
5、保留杭州信息，大部分新闻会点出“杭州+某地”比较好辨认，部分新闻可能会直接提到杭州具体的地址，比如闻涛路、武林广场等，大家稍作留意 ： keywords: 杭州、杭州地名。

6、如遇常规天气预报、车次信息、菜价播报等常规相关报道，可删除 ： keywords: 温度、雨、风
7、本次工作只需剔除非杭州信息，保留杭州新闻即可，不需要辨认具体地址也不需要打标等其他工作 ：

优先级最高的是：4、非杭州新闻内容，北京讯、南京讯等直接可删除，点明湖州、温州等其他地址的稿件一般情况都可直接剔除 : keywords: 全国城市名称。

'''

import openpyxl
import pymysql


def news_input(path, newss, newss_content, newss_date, newss_length, info=None):
    xls = openpyxl.load_workbook(path)
    sheet = xls["mlf_product_formal_20160511"]  # 'Sheet1'
    # xls = pd.read_excel(,sheetname="mlf_product_formal_20160511")

    rownum_of_sheet = sheet.max_row
    columnnum_of_sheet = sheet.max_column
    print("The sheet contains "+ str(rownum_of_sheet)+" rows, and "+str(columnnum_of_sheet)+" columns.")

    for i in range(2, rownum_of_sheet):
        news_date = sheet.cell(row=i, column=1).value
        news_content = sheet.cell(row=i, column=2).value
        news = news_date + news_content

        newss.append(news)
        newss_content.append(news_content)
        newss_date.append(news_date)
        newss_length.append(len(news_content))


def city_nationwide(path):
    cityfile = open(path)
    names = cityfile.readlines()
    cities = []

    for city in names:
        if '\n' in city:
            city = city.replace('\n', '')
        if '\r' in city:
            city = city.replace('\r', '')
        if city != '':
            if '市' in city:
                city = city.strip('市')
            if '省' in city:
                city = city.strip('省')
            cities.append(city)
    print(cities)
    return cities


def locations_hangzhou(path=None):
    localities = ['杭州', '西湖', '武林', '萧山']
    return localities


def store_news(newss_date, newss_content, filename):
    xlsx29 = openpyxl.Workbook()
    newsheet29 = xlsx29.active

    for i in range(2, len((newss_date))):
        news_date = newss_date[i]
        news_content= newss_content[i]

        newsheet29.cell(row=i, column=1, value=news_date)   #.value = newss_date
        newsheet29.cell(row=i, column=2, value=news_content)
        # print(news_date, news_content)

    xlsx29.save(filename)


# database login:
'''
connect = pymysql.connect(host='127.0.0.1', unix_socket='/tmp/mysql.sock', user='root', passwd='qibangdatabase2018', db='mysql')
sql_cursor = connect.cursor()
sql_cursor.execute("show databases")
print(sql_cursor.fetchall())
sql_cursor.execute("USE media_convergence")
print(sql_cursor.fetchall())
'''

# load data:
cities_nationwide = city_nationwide("city.txt")  # /Users/treasersmac/ZJU-ALL/_6_grand3_2/Media_/
hangzhou_localities = locations_hangzhou()
newss = []
newss_content = []
newss_date = []
newss_length = []
news_input("29.xlsx", newss, newss_content, newss_date, newss_length)  # /Users/treasersmac/ZJU-ALL/_6_grand3_2/Media_/6/
# show average words of these news:
average = 0
max = 0
for news_length in newss_length:
    average += news_length
    if news_length > max:
        max = news_length
average = average/len(newss_length)
print(max)  # 7610 9192 6031 9500 7147 6840
print(average)
# process data:
short_news_num = 0
outside_news_num = 0
l = len(newss_content)
Newss_Content = newss_content.copy()
Newss_Date = newss_date.copy()
Newss = newss.copy()
print(len(newss_date[1]))

for i in range(2, l):
    news_content = Newss_Content[i]
    news_date = Newss_Date[i]
    news = Newss[i]

    if 0 < len(news_content) < 80:          # 1、空信息，整行删除 : len(news) < 50
        short_news_num += 1
        if news_content in newss_content:
            newss_content.remove(news_content)   # operate raw data only
            newss_date.remove(news_date)         # operate raw data only

    for city_outside in cities_nationwide:  # 4、非杭州新闻内容，北京、南京，点明湖州、温州等地址的稿件直接剔除 : keywords: 全国城市名称。
        if city_outside in news_content:
            for hangzhou_locality in hangzhou_localities:
                if news_content in newss_content and hangzhou_locality not in newss_content:   # 5、保留杭州信息.
                    newss_content.remove(news_content)
                    newss_date.remove(news_date)
                    outside_news_num += 1
                break

print("short news num: " + str(short_news_num))
print("outside news num: " + str(outside_news_num))
# write data to new excel file:
# store_news(newss_date, newss_content, 'new5701.xlsx')




